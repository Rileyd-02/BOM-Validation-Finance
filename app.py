# app.py
import streamlit as st
import pandas as pd
from thefuzz import fuzz, process
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------
# Helper functions
# ---------------------------
def find_col(df, candidates, fallback_index=None):
    """Find first matching column name from candidates; fallback to index if provided."""
    for c in candidates:
        if c in df.columns:
            return c
    if fallback_index is not None and fallback_index < len(df.columns):
        return df.columns[fallback_index]
    return None

def safe_number(x):
    if x is None or x == "":
        return None
    try:
        return float(str(x).replace(",", ""))
    except:
        return None

def consumption_similarity(a, b):
    if a is None and b is None:
        return 100.0
    a = a or 0.0
    b = b or 0.0
    if a == 0 and b == 0:
        return 100.0
    denom = max(abs(a), abs(b), 1)
    pct_diff = abs(a - b) / denom * 100
    score = max(0.0, 100.0 - pct_diff)
    return round(score, 2)

def add_summary_sheet(wb, summary_dict):
    """Add a Summary sheet to workbook (openpyxl workbook)."""
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
    else:
        ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = True
    ws["A1"] = "Metric"
    ws["B1"] = "Count"
    row = 2
    for k, v in summary_dict.items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

# ---------------------------
# Streamlit UI: Upload & Map
# ---------------------------
st.set_page_config(layout="wide", page_title="SAP baseline → PLM validation (v2)")
st.title("📋 SAP baseline → PLM validation (v2)")
st.markdown("Upload SAP (baseline) and PLM files, confirm column mapping, set fuzzy threshold, then run comparison.")

col1, col2 = st.columns(2)
with col1:
    sap_file = st.file_uploader("Upload SAP file (Excel)", type=["xlsx", "xls"])
with col2:
    plm_file = st.file_uploader("Upload PLM file (Excel)", type=["xlsx", "xls"])

if not (sap_file and plm_file):
    st.info("Upload both SAP and PLM files to proceed.")
    st.stop()

# Read files (first sheet by default) and show detected columns
try:
    sap_df = pd.read_excel(sap_file, sheet_name=0, dtype=str)
    plm_df = pd.read_excel(plm_file, sheet_name=0, dtype=str)
except Exception as e:
    st.error(f"Error reading files: {e}")
    st.stop()

sap_df.columns = sap_df.columns.str.strip()
plm_df.columns = plm_df.columns.str.strip()

st.subheader("Detected columns — confirm or remap")
st.markdown("If a detected column is not correct, choose the correct column from the dropdown.")

# Heuristics for default mapping (using indexes you specified for fallback)
# SAP fallbacks: Material col C -> idx 2, Component G -> idx 6, MatDesc U -> idx 20, Vendor V -> idx 21
# PLM fallbacks: Material E -> idx 4, Vendor N -> idx 13
sap_material_default = find_col(sap_df, ["Material", "Material No", "Material Number", "Material Code"], fallback_index=2)
sap_component_default = find_col(sap_df, ["Component", "Component Code", "Component Name"], fallback_index=6)
sap_matdesc_default = find_col(sap_df, ["Material Description", "Material Desc", "Description"], fallback_index=20)
sap_vendor_default = find_col(sap_df, ["Vendor Reference", "Vendor Ref", "Vendor"], fallback_index=21)
sap_consumption_default = find_col(sap_df, ["Comp.Qty.", "Comp.Qty", "Consumption", "Qty(Cons.)", "Qty"], fallback_index=None)
sap_color_default = find_col(sap_df, ["Color Reference", "Color", "Color Name"], fallback_index=None)

plm_material_default = find_col(plm_df, ["Material", "Material No", "Material Number", "Material Code"], fallback_index=4)
plm_vendor_default = find_col(plm_df, ["Vendor Reference", "Vendor Ref", "Vendor"], fallback_index=13)
plm_consumption_default = find_col(plm_df, ["Qty(Cons.)", "Consumption", "Qty"], fallback_index=None)
plm_color_default = find_col(plm_df, ["Color Reference", "Color", "Color Name"], fallback_index=None)

# Provide mapping UI
with st.form("mapping_form"):
    st.write("**SAP -> select column**")
    sap_material_col = st.selectbox("SAP: Material (col C)", options=sap_df.columns.tolist(), index=sap_df.columns.get_loc(sap_material_default) if sap_material_default in sap_df.columns else 0)
    sap_component_col = st.selectbox("SAP: Component (col G)", options=sap_df.columns.tolist(), index=sap_df.columns.get_loc(sap_component_default) if sap_component_default in sap_df.columns else 0)
    sap_matdesc_col = st.selectbox("SAP: Material Description (col U)", options=sap_df.columns.tolist(), index=sap_df.columns.get_loc(sap_matdesc_default) if sap_matdesc_default in sap_df.columns else 0)
    sap_vendor_col = st.selectbox("SAP: Vendor Reference (col V)", options=sap_df.columns.tolist(), index=sap_df.columns.get_loc(sap_vendor_default) if sap_vendor_default in sap_df.columns else 0)
    sap_consumption_col = st.selectbox("SAP: Consumption column", options=sap_df.columns.tolist(), index=sap_df.columns.get_loc(sap_consumption_default) if sap_consumption_default in sap_df.columns else 0)
    sap_color_col = st.selectbox("SAP: Color column (optional)", options=[""] + sap_df.columns.tolist(), index=0 if not sap_color_default else (sap_df.columns.get_loc(sap_color_default)+1))

    st.write("**PLM -> select column**")
    plm_material_col = st.selectbox("PLM: Material (col E)", options=plm_df.columns.tolist(), index=plm_df.columns.get_loc(plm_material_default) if plm_material_default in plm_df.columns else 0)
    plm_vendor_col = st.selectbox("PLM: Vendor Reference (col N)", options=plm_df.columns.tolist(), index=plm_df.columns.get_loc(plm_vendor_default) if plm_vendor_default in plm_df.columns else 0)
    plm_consumption_col = st.selectbox("PLM: Consumption column", options=plm_df.columns.tolist(), index=plm_df.columns.get_loc(plm_consumption_default) if plm_consumption_default in plm_df.columns else 0)
    plm_color_col = st.selectbox("PLM: Color column (optional)", options=[""] + plm_df.columns.tolist(), index=0 if not plm_color_default else (plm_df.columns.get_loc(plm_color_default)+1))

    st.write("**Matching options**")
    fuzzy_threshold = st.slider("Material fuzzy-match threshold (for fallback auto-match)", min_value=60, max_value=100, value=85)
    submit_map = st.form_submit_button("Save mapping & continue")

if not submit_map:
    st.stop()

# show mapping recap
st.write("### Confirmed mapping")
mapping = {
    "SAP Material": sap_material_col,
    "SAP Component": sap_component_col,
    "SAP Material Description": sap_matdesc_col,
    "SAP Vendor": sap_vendor_col,
    "SAP Consumption": sap_consumption_col,
    "SAP Color": sap_color_col or "(none)",
    "PLM Material": plm_material_col,
    "PLM Vendor": plm_vendor_col,
    "PLM Consumption": plm_consumption_col,
    "PLM Color": plm_color_col or "(none)"
}
st.table(pd.DataFrame(list(mapping.items()), columns=["Field", "Mapped Column"]))

# Run comparison button
if st.button("Run comparison"):
    with st.spinner("Running comparison..."):
        # Normalize strings
        sap_df = sap_df.fillna("")
        plm_df = plm_df.fillna("")
        for df in (sap_df, plm_df):
            for c in df.columns:
                df[c] = df[c].astype(str).str.strip()

        # Build PLM index by material for exact lookup
        plm_index_by_material = {}
        for idx, prow in plm_df.iterrows():
            key = prow.get(plm_material_col, "")
            plm_index_by_material.setdefault(key, []).append((idx, prow))

        plm_material_values = plm_df[plm_material_col].tolist()

        results = []
        missing_in_plm_rows = []
        # iterate SAP baseline rows
        for sidx, srow in sap_df.iterrows():
            row_out = {}
            # base SAP fields
            sap_mat = srow.get(sap_material_col, "")
            sap_comp = srow.get(sap_component_col, "")
            sap_desc = srow.get(sap_matdesc_col, "")
            sap_vendor = srow.get(sap_vendor_col, "")
            sap_color = srow.get(sap_color_col, "")
            sap_cons_raw = srow.get(sap_consumption_col, "")
            sap_cons = safe_number(sap_cons_raw)

            # initialize output fields
            row_out["Material_SAP"] = sap_mat
            row_out["Material_Description_SAP"] = sap_desc
            row_out["Component_SAP"] = sap_comp
            row_out["Vendor_SAP"] = sap_vendor
            row_out["Color_SAP"] = sap_color
            row_out["SAP_Consumption"] = sap_cons
            row_out["Found_in_PLM"] = False
            row_out["Component_Flag"] = ""
            row_out["Vendor_Check"] = ""
            row_out["PLM_Material"] = ""
            row_out["Vendor_PLM"] = ""
            row_out["Color_PLM"] = ""
            row_out["PLM_Consumption"] = None
            row_out["ConsumptionDiff"] = None
            row_out["DifferenceFlag"] = ""
            row_out["Material_Similarity"] = 0
            row_out["Color_Similarity"] = 0
            row_out["Consumption_Similarity"] = 0
            row_out["Notes"] = ""

            # 1) Component check: if '-' in value or startswith '3' -> mark excluded (YELLOW), skip further checks
            if sap_comp and ("-" in sap_comp or sap_comp.lstrip().startswith("3")):
                row_out["Component_Flag"] = "Component excluded (contains '-' or starts with '3')"
                row_out["Notes"] = "Component exclusion - skipped further checks"
                results.append(row_out)
                continue

            # 2) Material exact match in PLM
            matched_plm_row = None
            matched_plm_idx = None
            if sap_mat and sap_mat in plm_index_by_material:
                matched_plm_idx, matched_plm_row = plm_index_by_material[sap_mat][0]
                row_out["Found_in_PLM"] = True
            else:
                # fuzzy fallback search for best PLM material
                if plm_material_values:
                    best = process.extractOne(sap_mat, plm_material_values, scorer=fuzz.token_sort_ratio)
                    if best and best[1] >= fuzzy_threshold:
                        # find row index for that value (first occurrence)
                        pm_val = best[0]
                        pm_idx = plm_df[plm_df[plm_material_col] == pm_val].index
                        if len(pm_idx) > 0:
                            matched_plm_idx = pm_idx[0]
                            matched_plm_row = plm_df.loc[matched_plm_idx]
                            row_out["Found_in_PLM"] = True
                            row_out["Notes"] = f"Fuzzy material match (score {best[1]})"
                    else:
                        row_out["Notes"] = f"Material not found in PLM (best material fuzzy score {best[1] if best else 'N/A'})"
                        # record as missing and skip vendor/consumption checks
                        missing_in_plm_rows.append(row_out)
                        results.append(row_out)
                        continue
                else:
                    row_out["Notes"] = "PLM materials list empty"
                    missing_in_plm_rows.append(row_out)
                    results.append(row_out)
                    continue

            # if matched_plm_row exists, proceed vendor check
            if matched_plm_row is not None:
                plm_vendor_val = matched_plm_row.get(plm_vendor_col, "")
                plm_color_val = matched_plm_row.get(plm_color_col, "") if plm_color_col else ""
                plm_cons_raw = matched_plm_row.get(plm_consumption_col, "") if plm_consumption_col else ""
                plm_cons = safe_number(plm_cons_raw)

                row_out["PLM_Material"] = matched_plm_row.get(plm_material_col, "")
                row_out["Vendor_PLM"] = plm_vendor_val
                row_out["Color_PLM"] = plm_color_val
                row_out["PLM_Consumption"] = plm_cons

                # Vendor check: exact in SAP vendor column
                vendor_ok = False
                if sap_vendor and plm_vendor_val and sap_vendor == plm_vendor_val:
                    vendor_ok = True
                    row_out["Vendor_Check"] = "Vendor OK (exact match)"
                else:
                    # check if plm_vendor_val exists inside sap material description
                    if plm_vendor_val and plm_vendor_val in sap_desc:
                        vendor_ok = True
                        row_out["Vendor_Check"] = "Vendor OK (found in material description)"
                    else:
                        vendor_ok = False
                        row_out["Vendor_Check"] = "Vendor Not Found"
                        row_out["Notes"] = (row_out["Notes"] + " | Vendor not found in SAP vendor column or material description").strip(" |")

                # Color similarity
                if plm_color_val and sap_color:
                    row_out["Color_Similarity"] = fuzz.token_sort_ratio(sap_color.lower(), plm_color_val.lower())
                else:
                    row_out["Color_Similarity"] = 0

                # Consumption diff & flag
                if sap_cons is None or plm_cons is None:
                    row_out["ConsumptionDiff"] = None
                    row_out["DifferenceFlag"] = "Missing consumption value"
                else:
                    diff = plm_cons - sap_cons
                    row_out["ConsumptionDiff"] = diff
                    row_out["Consumption_Similarity"] = consumption_similarity(sap_cons, plm_cons)
                    if sap_cons > plm_cons:
                        row_out["DifferenceFlag"] = "SAP consumption is higher"
                    else:
                        row_out["DifferenceFlag"] = "OK"

                # Material similarity
                row_out["Material_Similarity"] = fuzz.token_sort_ratio(sap_mat.lower(), matched_plm_row.get(plm_material_col, "").lower()) if sap_mat and matched_plm_row.get(plm_material_col, "") else 0

                # If vendor not found, mark vendor mismatch but keep processing (we highlight later)
                results.append(row_out)
            else:
                # should not reach here, but guard
                row_out["Notes"] = "No PLM match row found (unexpected path)"
                results.append(row_out)

        # Build final results dataframe
        results_df = pd.DataFrame(results)

        # Prepare summary numbers
        total_rows = len(results_df)
        not_in_plm = results_df["Found_in_PLM"].eq(False).sum()
        comp_excluded = results_df["Component_Flag"].ne("").sum()
        vendor_mismatch = results_df["Vendor_Check"].eq("Vendor Not Found").sum()
        sap_higher = results_df["DifferenceFlag"].eq("SAP consumption is higher").sum()
        ok_count = results_df["DifferenceFlag"].eq("OK").sum()

        # Show summary in UI
        st.subheader("Summary")
        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Total SAP rows processed", total_rows)
        s2.metric("Not in PLM", int(not_in_plm))
        s3.metric("Component excluded (yellow)", int(comp_excluded))
        s4.metric("Vendor mismatch", int(vendor_mismatch))
        s5, s6 = st.columns(2)
        s5.metric("SAP consumption is higher", int(sap_higher))
        s6.metric("OK (PLM >= SAP)", int(ok_count))

        # Create Excel output with Results, Missing_in_PLM, SAP_raw, PLM_raw, Summary
        out_buffer = BytesIO()
        with pd.ExcelWriter(out_buffer, engine="openpyxl") as writer:
            results_df.to_excel(writer, sheet_name="Results", index=False)
            if missing_in_plm_rows:
                pd.DataFrame(missing_in_plm_rows).to_excel(writer, sheet_name="Missing_in_PLM", index=False)
            sap_df.to_excel(writer, sheet_name="SAP_raw", index=False)
            plm_df.to_excel(writer, sheet_name="PLM_raw", index=False)
            # summary sheet will be added post-save using openpyxl
        out_buffer.seek(0)

        # Load wb and apply partial-cell highlighting:
        wb = load_workbook(out_buffer)
        # Add Summary sheet
        summary_dict = {
            "Total SAP rows processed": total_rows,
            "Not in PLM": int(not_in_plm),
            "Component excluded": int(comp_excluded),
            "Vendor mismatch": int(vendor_mismatch),
            "SAP consumption is higher": int(sap_higher),
            "OK (PLM >= SAP)": int(ok_count)
        }
        add_summary_sheet(wb, summary_dict)

        # Get Results sheet and headers to find indices
        if "Results" in wb.sheetnames:
            ws = wb["Results"]
            headers = [cell.value for cell in ws[1]]
            # helper to get column index number
            def idx(name):
                return headers.index(name) + 1 if name in headers else None

            col_comp_flag = idx("Component_Flag")
            col_found = idx("Found_in_PLM")
            col_vendor_check = idx("Vendor_Check")
            col_difference_flag = idx("DifferenceFlag")
            col_vendor_plm = idx("Vendor_PLM")
            col_vendor_sap = idx("Vendor_SAP")
            col_sap_cons = idx("SAP_Consumption")
            col_plm_cons = idx("PLM_Consumption")
            col_color_plm = idx("Color_PLM")
            col_color_sap = idx("Color_SAP")

            # fills
            yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            orange = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            red = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            for r in range(2, ws.max_row + 1):
                # read flags
                comp_flag_val = ws.cell(row=r, column=col_comp_flag).value if col_comp_flag else ""
                found_val = ws.cell(row=r, column=col_found).value if col_found else True
                vendor_check_val = ws.cell(row=r, column=col_vendor_check).value if col_vendor_check else ""
                diff_flag_val = ws.cell(row=r, column=col_difference_flag).value if col_difference_flag else ""

                # priority coloring: component exclusion -> yellow (component cell only),
                # not found in PLM or vendor not found -> orange (vendor columns only),
                # consumption flags -> red/green (consumption cells only)
                # apply Component cell yellow
                if comp_flag_val and "Component excluded" in str(comp_flag_val):
                    if col_comp_flag:
                        ws.cell(row=r, column=col_comp_flag).fill = yellow
                # vendor mismatch / not in PLM -> highlight vendor cells (PLM and SAP vendor columns if present)
                if (found_val in [False, "False", "FALSE", "", 0]) or ("Vendor Not Found" in str(vendor_check_val)):
                    if col_vendor_plm:
                        ws.cell(row=r, column=col_vendor_plm).fill = orange
                    if col_vendor_sap:
                        ws.cell(row=r, column=col_vendor_sap).fill = orange
                # consumption color
                if diff_flag_val == "SAP consumption is higher":
                    if col_sap_cons:
                        ws.cell(row=r, column=col_sap_cons).fill = red
                    if col_plm_cons:
                        ws.cell(row=r, column=col_plm_cons).fill = red
                elif diff_flag_val == "OK":
                    if col_sap_cons:
                        ws.cell(row=r, column=col_sap_cons).fill = green
                    if col_plm_cons:
                        ws.cell(row=r, column=col_plm_cons).fill = green

                # option: highlight color mismatch columns (light) if needed (not applied by default)

        # finalize workbook to bytes
        final_bytes = BytesIO()
        wb.save(final_bytes)
        final_bytes.seek(0)

        # UI previews & download
        st.success("Comparison complete.")
        st.download_button("📥 Download full report (xlsx)", data=final_bytes, file_name="sap_plm_validation_v2.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.subheader("Results preview (selected columns)")
        preview_cols = ["Material_SAP", "Material_Description_SAP", "Vendor_PLM", "Vendor_SAP", "Color_PLM", "Color_SAP", "SAP_Consumption", "PLM_Consumption", "ConsumptionDiff", "DifferenceFlag", "Component_Flag", "Vendor_Check", "Notes"]
        preview_cols = [c for c in preview_cols if c in results_df.columns]
        st.dataframe(results_df[preview_cols].fillna("").head(500), use_container_width=True)

        st.subheader("Full results (first 200 rows)")
        st.dataframe(results_df.head(200), use_container_width=True)

        st.info("If mappings are incorrect for future runs, re-upload and adjust column mapping before running.")

else:
    st.stop()
